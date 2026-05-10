"""Parse uploaded files into text and image data for the generation pipeline."""

import base64
import io
import os
import tempfile
from pathlib import Path

from docx import Document
from openpyxl import load_workbook


def parse_uploaded_file(file_bytes: bytes, filename: str, video_max_frames: int = 8) -> dict:
    """Parse an uploaded file, returning {'type': 'text'|'image'|'video', 'text': str, 'images': list}.

    Video files are sampled into frames (images) for Claude vision.
    """
    ext = Path(filename).suffix.lower()
    if ext == ".docx":
        return _parse_docx(file_bytes)
    elif ext in (".xlsx", ".xls"):
        return _parse_xlsx(file_bytes)
    elif ext in (".png", ".jpg", ".jpeg", ".gif", ".webp"):
        return _parse_image(file_bytes, ext)
    elif ext in (".mp4", ".mov", ".avi", ".mkv", ".webm", ".flv", ".wmv"):
        return _parse_video(file_bytes, ext, video_max_frames)
    elif ext == ".txt":
        return {"type": "text", "text": file_bytes.decode("utf-8"), "images": []}
    else:
        return {"type": "text", "text": "", "images": []}


def _parse_docx(file_bytes: bytes) -> dict:
    doc = Document(io.BytesIO(file_bytes))

    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    text = "\n".join(paragraphs)

    images = []
    for rel in doc.part.rels.values():
        if "image" in rel.reltype:
            try:
                image_data = rel.target_part.blob
                ext = Path(rel.target_part.partname).suffix.lower()
                mime = _mime_type(ext)
                b64 = base64.b64encode(image_data).decode("utf-8")
                images.append({"data": b64, "media_type": mime})
            except Exception:
                pass

    return {"type": "text", "text": text, "images": images}


def _parse_xlsx(file_bytes: bytes) -> dict:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    parts = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        parts.append(f"--- Sheet: {name} ---")
        for row in rows:
            cells = [str(c) if c is not None else "" for c in row]
            parts.append(" | ".join(cells))
    wb.close()
    return {"type": "text", "text": "\n".join(parts), "images": []}


def _parse_image(file_bytes: bytes, ext: str) -> dict:
    b64 = base64.b64encode(file_bytes).decode("utf-8")
    return {
        "type": "image",
        "text": "",
        "images": [{"data": b64, "media_type": _mime_type(ext)}],
    }


def _parse_video(file_bytes: bytes, ext: str, max_frames: int = 8) -> dict:
    """Extract key frames from video as base64 images for Claude vision."""
    import cv2
    import numpy as np

    # Write video bytes to temp file (OpenCV needs a file path)
    tmp_path = None
    try:
        fd, tmp_path = tempfile.mkstemp(suffix=ext)
        os.write(fd, file_bytes)
        os.close(fd)

        cap = cv2.VideoCapture(tmp_path)
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        fps = cap.get(cv2.CAP_PROP_FPS)
        duration = total_frames / fps if fps > 0 else 0

        if total_frames == 0:
            cap.release()
            return {"type": "video", "text": "", "images": []}

        # Sample evenly spaced frames
        frame_count = min(max_frames, total_frames)
        if frame_count < total_frames:
            step = total_frames // (frame_count + 1)
        else:
            step = 1

        images = []
        for i in range(frame_count):
            frame_idx = step * (i + 1) - 1
            cap.set(cv2.CAP_PROP_POS_FRAMES, frame_idx)
            ret, frame = cap.read()
            if not ret:
                continue

            # Resize if too large (max 1568px on longest side, per Claude API limits)
            h, w = frame.shape[:2]
            max_side = 1568
            if max(h, w) > max_side:
                scale = max_side / max(h, w)
                new_w, new_h = int(w * scale), int(h * scale)
                frame = cv2.resize(frame, (new_w, new_h))

            # Encode as JPEG
            _, buf = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 80])
            b64 = base64.b64encode(buf.tobytes()).decode("utf-8")
            images.append({"data": b64, "media_type": "image/jpeg"})

        cap.release()

        info = f"[视频: {duration:.0f}秒, 从 {total_frames} 帧中提取 {len(images)} 帧]"
        return {"type": "video", "text": info, "images": images}

    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


def _mime_type(ext: str) -> str:
    return {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".gif": "image/gif",
        ".webp": "image/webp",
    }.get(ext, "image/png")
